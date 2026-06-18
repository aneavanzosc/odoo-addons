# Copyright 2026 Ane Gurruchaga - AvanzOSC
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl.html).
import base64
import io
import logging

import openpyxl

from odoo import _, api, fields, models

_logger = logging.getLogger(__name__)


class ProductPricelistItemImport(models.Model):
    _name = "product.pricelist.item.import"
    _inherit = "base.import"
    _description = "Wizard to import pricelist items"

    import_line_ids = fields.One2many(
        comodel_name="product.pricelist.item.import.line",
    )
    catalog_ids = fields.Many2many(
        comodel_name="product.catalog",
        string="Catalogs",
    )

    def _read_xlsx(self):
        workbook = openpyxl.load_workbook(
            io.BytesIO(base64.decodebytes(self.data)), data_only=True
        )
        sheet = (
            workbook["Products"]
            if "Products" in workbook.sheetnames
            else workbook.active
        )
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        headers = [self._clean(header) for header in header_cells]
        pricelists = self.env["product.pricelist"].search([("market_id", "!=", False)])
        lines = []
        for pricelist in pricelists:
            columns = self._get_shopify_columns(headers, pricelist.market_id)
            if not columns:
                continue
            lines += self._get_shopify_line_values(sheet, columns, pricelist)
        _logger.info("Pricelist item import generated %s lines.", len(lines))
        return lines

    _read_xls = _read_xlsx

    def _get_shopify_line_values(self, sheet, columns, pricelist):
        lines = []
        for row_number, cells in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            price = self._to_optional_float(self._cell_value(cells, columns["price"]))
            compare_price = self._to_optional_float(
                self._cell_value(cells, columns["compare_price"])
            )
            if price is False and compare_price is False:
                continue
            if columns.get("included") is not None and not self._to_bool(
                self._cell_value(cells, columns["included"]), default=True
            ):
                continue
            identifiers = self._get_shopify_identifiers(cells, columns)
            if not identifiers["variant_sku"] and not identifiers["variant_barcode"]:
                continue
            prices = self._get_shopify_prices(price, compare_price)
            lines.append(
                (
                    0,
                    0,
                    {
                        "import_id": self.id,
                        "row_number": row_number,
                        "pricelist_id": pricelist.id,
                        "market_id": pricelist.market_id.id,
                        "catalog_ids": [(6, 0, self.catalog_ids.ids)],
                        "variant_sku": identifiers["variant_sku"],
                        "variant_barcode": identifiers["variant_barcode"],
                        "price": price,
                        "compare_price": compare_price,
                        "fixed_price": prices["fixed_price"],
                        "distribution_price": prices["distribution_price"],
                    },
                )
            )
        return lines

    @staticmethod
    def _get_shopify_prices(price, compare_price):
        if compare_price is not False:
            return {
                "fixed_price": compare_price,
                "distribution_price": price,
            }
        return {
            "fixed_price": price,
            "distribution_price": price,
        }

    def _get_shopify_columns(self, headers, market):
        normalized_market = self._normalize_header(market.name)
        price_col = compare_price_col = included_col = None
        for index, header in enumerate(headers):
            if "/" not in header:
                continue
            concept, header_market = self._split_shopify_header(header)
            if header_market != normalized_market:
                continue
            if concept == "price":
                price_col = index
            elif concept == "compare at price":
                compare_price_col = index
            elif concept == "included":
                included_col = index
        if price_col is None:
            return {}
        return {
            "price": price_col,
            "compare_price": compare_price_col,
            "included": included_col,
            "identifier": self._get_shopify_identifier_columns(headers),
        }

    def _split_shopify_header(self, header):
        concept, header_market = header.split("/", 1)
        return self._normalize_header(concept), self._normalize_header(header_market)

    def _get_shopify_identifier_columns(self, headers):
        candidates = {
            "variant_sku": ("variant sku", "sku", "referencia interna"),
            "variant_barcode": ("variant barcode", "barcode", "codigo de barras"),
        }
        result = {}
        for index, header in enumerate(headers):
            normalized_header = self._normalize_header(header)
            for field_name, aliases in candidates.items():
                if normalized_header in aliases:
                    result[field_name] = index
        return result

    def _get_shopify_identifiers(self, cells, columns):
        identifier_columns = columns["identifier"]
        return {
            "variant_sku": self._clean_cell(
                cells, identifier_columns.get("variant_sku")
            ),
            "variant_barcode": self._clean_cell(
                cells, identifier_columns.get("variant_barcode")
            ),
        }

    @staticmethod
    def _clean(value):
        if value is None:
            return ""
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @classmethod
    def _clean_cell(cls, cells, index):
        return cls._clean(cls._cell_value(cells, index))

    @staticmethod
    def _cell_value(cells, index):
        if index is None or index >= len(cells):
            return ""
        return cells[index]

    @staticmethod
    def _normalize_header(value):
        replacements = str.maketrans("áéíóúüñ", "aeiouun")
        value = str(value or "").replace("\xa0", " ")
        return " ".join(value.strip().lower().translate(replacements).split())

    @staticmethod
    def _to_optional_float(value):
        if value in (None, ""):
            return False
        value = ProductPricelistItemImport._normalize_number(value)
        return float(value) if value not in (None, "") else False

    @staticmethod
    def _normalize_number(value):
        if not isinstance(value, str):
            return value
        value = value.strip().replace("$", "").replace("€", "").replace("£", "")
        value = value.replace(" ", "")
        if "," in value and "." in value:
            value = value.replace(".", "").replace(",", ".")
        else:
            value = value.replace(",", ".")
        return value

    @staticmethod
    def _to_bool(value, default=False):
        if value in (None, ""):
            return default
        if isinstance(value, str):
            return value.strip().lower() in (
                "1",
                "true",
                "t",
                "yes",
                "y",
                "si",
                "verdadero",
            )
        return bool(value)


class ProductPricelistItemImportLine(models.Model):
    _name = "product.pricelist.item.import.line"
    _inherit = "base.import.line"
    _description = "Lines to import pricelist items"

    import_id = fields.Many2one(
        comodel_name="product.pricelist.item.import",
        string="Import Wizard",
        copy=False,
    )
    action = fields.Selection(
        selection_add=[
            ("create", "Create"),
            ("update", "Update"),
        ],
        ondelete={
            "create": "set default",
            "update": "set default",
        },
    )
    row_number = fields.Integer(
        copy=False,
        readonly=True,
    )
    pricelist_id = fields.Many2one(
        comodel_name="product.pricelist",
        string="Pricelist",
        copy=False,
        readonly=True,
    )
    market_id = fields.Many2one(
        comodel_name="product.pricelist.market",
        string="Market",
        copy=False,
        readonly=True,
    )
    catalog_ids = fields.Many2many(
        comodel_name="product.catalog",
        string="Catalogs",
        copy=False,
        readonly=True,
    )
    variant_sku = fields.Char(
        copy=False,
        readonly=True,
    )
    variant_barcode = fields.Char(
        copy=False,
        readonly=True,
    )
    product_id = fields.Many2one(
        comodel_name="product.product",
        string="Product",
        copy=False,
        readonly=True,
    )
    pricelist_item_ids = fields.Many2many(
        comodel_name="product.pricelist.item",
        string="Pricelist Items",
        copy=False,
        readonly=True,
    )
    price = fields.Float(
        copy=False,
    )
    compare_price = fields.Float(
        copy=False,
    )
    fixed_price = fields.Float(
        copy=False,
    )
    distribution_price = fields.Float(
        copy=False,
    )

    @api.onchange("price", "compare_price")
    def _onchange_prices(self):
        for line in self:
            line.fixed_price = line.compare_price or line.price
            line.distribution_price = line.price

    def _action_validate(self):
        update_values = super()._action_validate()
        log_infos = []
        product = self._find_product()
        current_items = self.env["product.pricelist.item"]
        if not product:
            log_infos.append(_("Error: Product not found."))
        else:
            current_items = self._find_current_pricelist_items(product)
        state = "error" if log_infos else "pass"
        action = "nothing"
        if state == "pass":
            action = "update" if current_items else "create"
        update_values.update(
            {
                "product_id": product and product.id,
                "pricelist_item_ids": [(6, 0, current_items.ids)],
                "log_info": "\n".join(log_infos),
                "state": state,
                "action": action,
            }
        )
        return update_values

    def _action_process(self):
        update_values = super()._action_process()
        log_info = ""
        if self.action == "update":
            self.pricelist_item_ids.write(self._pricelist_item_values())
        elif self.action == "create":
            product = self.product_id or self._find_product()
            if product:
                self.env["product.pricelist.item"].create(
                    self._pricelist_item_values(product=product, create=True)
                )
            else:
                log_info = _("Error: Product not found.")
        update_values.update(
            {
                "state": "error" if log_info else "done",
                "log_info": log_info,
            }
        )
        return update_values

    def _find_product(self):
        self.ensure_one()
        product_obj = self.env["product.product"]
        if self.product_id:
            return self.product_id
        if self.variant_sku:
            product = product_obj.search([("default_code", "=", self.variant_sku)])
            if len(product) == 1:
                return product
        if self.variant_barcode:
            product = product_obj.search([("barcode", "=", self.variant_barcode)])
            if len(product) == 1:
                return product
        return product_obj

    def _find_current_pricelist_items(self, product):
        self.ensure_one()
        items = self.env["product.pricelist.item"].search(
            [
                ("pricelist_id", "=", self.pricelist_id.id),
                ("applied_on", "=", "0_product_variant"),
                ("product_id", "=", product.id),
                ("compute_price", "=", "fixed"),
            ]
        )
        today = fields.Date.context_today(self)
        return items.filtered(
            lambda item: not item.date_end
            or fields.Date.to_date(item.date_end) >= today
        )

    def _pricelist_item_values(self, product=False, create=False):
        self.ensure_one()
        values = {
            "fixed_price": self.fixed_price,
            "distribution_price": self.distribution_price,
            "catalog_ids": [(6, 0, self.catalog_ids.ids)],
        }
        if create:
            values.update(
                {
                    "pricelist_id": self.pricelist_id.id,
                    "applied_on": "0_product_variant",
                    "product_id": product.id,
                    "min_quantity": 1.0,
                    "compute_price": "fixed",
                }
            )
        return values
